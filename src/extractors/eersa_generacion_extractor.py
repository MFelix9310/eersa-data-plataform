"""
Extractor Bronze: xlsx de generación EERSA 2021 → Parquet particionado.
Lee los 12 archivos mensuales, hace unpivot wide→long, agrega metadatos Bronze.
"""

import logging
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# Raíz del proyecto (2 niveles arriba de este archivo)
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
RAW_DIR = PROJECT_ROOT / "data" / "raw"
BRONZE_DIR = PROJECT_ROOT / "data" / "bronze" / "eersa_generacion"

# Mapeo de columnas por planta (1-indexed como openpyxl)
# Plantas con 4 métricas: KW, E.Bruta kWh, C.Int kWh, E.Neta kWh
PLANTAS_4_METRICAS = {
    "ALAO": {"KW": 3, "E.Bruta kWh": 4, "C.Int kWh": 5, "E.Neta kWh": 6},
    "RIO BLANCO": {"KW": 7, "E.Bruta kWh": 8, "C.Int kWh": 9, "E.Neta kWh": 10},
    "C.NIZAG": {"KW": 11, "E.Bruta kWh": 12, "C.Int kWh": 13, "E.Neta kWh": 14},
}

# S.N.I. solo tiene 2 métricas
PLANTAS_2_METRICAS = {
    "S.N.I.": {"KW": 15, "KWh": 16},
}

# Totales y MEM
AGREGADOS = {
    "TOTAL EERSA": {"KW": 20, "KW-H": 21},
    "ENERGIA ENTREGADA AL MEM": {"kWh": 22},
    "ENERGIA RECIBIDA DEL MEM": {"kWh": 23},
}

# Meses en español → número
MESES_ES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

HEADER_ROW = 9
DATA_START_ROW = 11


def extraer_mes_anio(ws) -> tuple[int, int]:
    """Extrae mes y año desde la fila 5 del Excel."""
    mes_str = str(ws.cell(row=5, column=3).value).strip().upper()
    anio = int(ws.cell(row=5, column=5).value)
    mes = MESES_ES.get(mes_str)
    if mes is None:
        raise ValueError(f"Mes no reconocido: '{mes_str}'")
    return mes, anio


def extraer_datos_archivo(filepath: Path, batch_id: str) -> pd.DataFrame:
    """Lee un xlsx y devuelve DataFrame long (fecha, planta, metrica, valor)."""
    logger.info(f"Procesando: {filepath.name}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb["RESUMEN"]
    mes, anio = extraer_mes_anio(ws)

    registros = []
    ingested_at = datetime.now(timezone.utc).isoformat()

    row = DATA_START_ROW
    while True:
        dia_val = ws.cell(row=row, column=2).value
        if dia_val is None or str(dia_val).strip() == "":
            row += 1
            # Limite de seguridad
            if row > 80:
                break
            continue

        dia_str = str(dia_val).strip()
        if dia_str.upper() == "TOTAL":
            break

        try:
            dia = int(float(dia_str))
        except (ValueError, TypeError):
            row += 1
            continue

        try:
            fecha = datetime(anio, mes, dia).date()
        except ValueError:
            logger.warning(f"Fecha inválida: {anio}-{mes}-{dia} en {filepath.name}")
            row += 1
            continue

        # Plantas con 4 métricas
        for planta, metricas in PLANTAS_4_METRICAS.items():
            for metrica, col in metricas.items():
                valor = ws.cell(row=row, column=col).value
                registros.append({
                    "fecha": fecha,
                    "planta": planta,
                    "metrica": metrica,
                    "valor": _limpiar_valor(valor),
                    "mes": mes,
                    "anio": anio,
                    "_source_file": filepath.name,
                    "_ingested_at": ingested_at,
                    "_batch_id": batch_id,
                })

        # Plantas con 2 métricas
        for planta, metricas in PLANTAS_2_METRICAS.items():
            for metrica, col in metricas.items():
                valor = ws.cell(row=row, column=col).value
                registros.append({
                    "fecha": fecha,
                    "planta": planta,
                    "metrica": metrica,
                    "valor": _limpiar_valor(valor),
                    "mes": mes,
                    "anio": anio,
                    "_source_file": filepath.name,
                    "_ingested_at": ingested_at,
                    "_batch_id": batch_id,
                })

        # Agregados (TOTAL EERSA, MEM)
        for planta, metricas in AGREGADOS.items():
            for metrica, col in metricas.items():
                valor = ws.cell(row=row, column=col).value
                registros.append({
                    "fecha": fecha,
                    "planta": planta,
                    "metrica": metrica,
                    "valor": _limpiar_valor(valor),
                    "mes": mes,
                    "anio": anio,
                    "_source_file": filepath.name,
                    "_ingested_at": ingested_at,
                    "_batch_id": batch_id,
                })

        row += 1

    wb.close()

    if not registros:
        logger.warning(f"Sin datos extraídos de {filepath.name}")
        return pd.DataFrame()

    df = pd.DataFrame(registros)
    df["fecha"] = pd.to_datetime(df["fecha"])
    logger.info(f"  → {len(df)} registros, {df['planta'].nunique()} plantas, días {df['fecha'].dt.day.min()}-{df['fecha'].dt.day.max()}")
    return df


def _limpiar_valor(valor) -> float | None:
    """Convierte valor de celda a float, None si vacío/no numérico."""
    if valor is None or str(valor).strip() in ("", "dm", "-"):
        return None
    try:
        return float(valor)
    except (ValueError, TypeError):
        return None


def guardar_parquet(df: pd.DataFrame, anio: int, mes: int):
    """Guarda DataFrame como Parquet particionado por mes."""
    out_dir = BRONZE_DIR / f"year={anio}" / f"month={mes:02d}"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "data.parquet"

    table = pa.Table.from_pandas(df, preserve_index=False)
    pq.write_table(table, out_path, compression="snappy")
    logger.info(f"  → Guardado: {out_path} ({len(df)} filas)")


def run():
    """Ejecuta la ingesta completa Bronze."""
    logger.info("=" * 60)
    logger.info("INGESTA BRONZE - Generación EERSA 2021")
    logger.info("=" * 60)

    archivos = sorted(RAW_DIR.glob("*.xlsx"))
    if not archivos:
        logger.error(f"No se encontraron archivos .xlsx en {RAW_DIR}")
        return

    batch_id = str(uuid.uuid4())
    logger.info(f"Batch ID: {batch_id}")
    logger.info(f"Archivos encontrados: {len(archivos)}")

    total_filas = 0
    todas_plantas = set()

    for filepath in archivos:
        try:
            df = extraer_datos_archivo(filepath, batch_id)
            if df.empty:
                continue

            mes = df["mes"].iloc[0]
            anio = df["anio"].iloc[0]
            guardar_parquet(df, anio, mes)

            total_filas += len(df)
            todas_plantas.update(df["planta"].unique())

        except Exception as e:
            logger.error(f"Error procesando {filepath.name}: {e}", exc_info=True)

    logger.info("=" * 60)
    logger.info("RESUMEN INGESTA BRONZE")
    logger.info(f"  Archivos procesados: {len(archivos)}")
    logger.info(f"  Filas totales: {total_filas:,}")
    logger.info(f"  Plantas detectadas: {sorted(todas_plantas)}")
    logger.info(f"  Output: {BRONZE_DIR}")
    logger.info("=" * 60)


if __name__ == "__main__":
    run()
